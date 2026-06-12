# app.py — Combinix Lovelace · CSP Solver: blocos 2/3 + cascata níveis + sábado 3 estados
import os, sys, json, copy, io, logging, uuid, secrets, hmac

# A edição local inclui dependências Python portáteis em ``vendor/``.
# Assim, o sistema abre sem executar pip install e sem depender de internet.
_APP_ROOT = os.path.dirname(os.path.abspath(__file__))
_VENDOR_DIR = os.path.join(_APP_ROOT, 'vendor')
if os.path.isdir(_VENDOR_DIR) and _VENDOR_DIR not in sys.path:
    sys.path.insert(0, _VENDOR_DIR)

from collections import defaultdict
from flask import Flask, render_template, request, session, jsonify, send_file, g
from werkzeug.exceptions import RequestEntityTooLarge
from persistence import save_state, load_state, reset_state, normalize_workspace_id
from services.solver import (
    solve_schedule, disc_uid as solver_disc_uid, grupo_key as solver_grupo_key,
    cfg_padrao as solver_cfg_padrao, cfg_padrao_prof as solver_cfg_padrao_prof,
    normalizar_avancadas as solver_normalizar_avancadas,
    get_dias_para_nivel as solver_get_dias_para_nivel,
    analisar_cobertura_professores as solver_analisar_cobertura_professores,
)

APP_VERSION = '2.2.2-hybrid'

# Persistência híbrida automática:
# - local: JSONs dentro da pasta do projeto;
# - browser: IndexedDB no navegador (ativado automaticamente na Vercel).
STORAGE_MODE = 'browser' if (
    os.environ.get('COMBINIX_STORAGE_MODE', '').strip().lower() == 'browser'
    or bool(os.environ.get('VERCEL'))
) else 'local'
# O estado web cruza uma Vercel Function em cada operação. Mantemos margem
# abaixo do limite de payload da hospedagem para comportar o envelope JSON.
BROWSER_STATE_MAX_BYTES = 3 * 1024 * 1024
_BROWSER_RESULTS_KEY = 'browser_resultados'

app = Flask(__name__)

def _load_or_create_secret_key():
    # Em produção, prefira COMBINIX_SECRET_KEY estável. No modo navegador a
    # proteção CSRF usa cookie próprio e não depende da persistência em disco.
    env_secret = os.environ.get('COMBINIX_SECRET_KEY', '').strip()
    if env_secret:
        return env_secret
    if STORAGE_MODE == 'browser':
        return secrets.token_urlsafe(48)
    db_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'database')
    os.makedirs(db_dir, exist_ok=True)
    path = os.path.join(db_dir, '.secret_key')
    try:
        if os.path.exists(path):
            with open(path, 'r', encoding='utf-8') as handle:
                value = handle.read().strip()
            if value:
                return value
        value = secrets.token_urlsafe(48)
        with open(path, 'w', encoding='utf-8') as handle:
            handle.write(value)
        return value
    except OSError:
        logging.warning('[Combinix] Não foi possível persistir .secret_key; usando segredo temporário.')
        return secrets.token_urlsafe(48)

app.secret_key = _load_or_create_secret_key()
app.config['MAX_CONTENT_LENGTH'] = (4 if STORAGE_MODE == 'browser' else 16) * 1024 * 1024  # backups: margem segura na Vercel

def _workspace_id():
    # No modo navegador, o workspace real mora no IndexedDB daquela origem.
    # Após login, este identificador poderá ser substituído pelo ID do usuário.
    default = 'browser' if STORAGE_MODE == 'browser' else 'local'
    workspace = normalize_workspace_id(session.get('workspace_id', default))
    session['workspace_id'] = workspace
    return workspace


def _csrf_token():
    """Token CSRF estável também em runtimes serverless.

    No modo navegador usamos double-submit cookie: o valor enviado pelo HTML é
    comparado com o cookie do próprio domínio. Assim, um novo processo da Vercel
    não invalida a página aberta. No modo local preservamos a sessão Flask.
    """
    if STORAGE_MODE == 'browser':
        token = getattr(g, '_combinix_csrf_token', None)
        if token:
            return token
        token = request.cookies.get('combinix_csrf', '').strip() or secrets.token_urlsafe(32)
        g._combinix_csrf_token = token
        return token
    token = session.get('_csrf_token')
    if not token:
        token = secrets.token_urlsafe(32)
        session['_csrf_token'] = token
    return token


def _decode_browser_state(raw):
    if raw in (None, ''):
        return {}
    if isinstance(raw, str):
        if len(raw.encode('utf-8')) > BROWSER_STATE_MAX_BYTES:
            raise ValueError('Estado do navegador acima do limite permitido.')
        try:
            raw = json.loads(raw)
        except json.JSONDecodeError as exc:
            raise ValueError('Estado do navegador inválido.') from exc
    if not isinstance(raw, dict):
        raise ValueError('Estado do navegador inválido.')
    try:
        encoded = json.dumps(raw, ensure_ascii=False).encode('utf-8')
    except (TypeError, ValueError) as exc:
        raise ValueError('Estado do navegador inválido.') from exc
    if len(encoded) > BROWSER_STATE_MAX_BYTES:
        raise ValueError('Estado do navegador acima do limite permitido.')
    return raw


def _browser_state_from_request():
    raw = None
    if request.is_json:
        payload = request.get_json(silent=True) or {}
        if isinstance(payload, dict):
            raw = payload.get('_browser_state')
    if raw is None:
        raw = request.form.get('browser_state') or request.form.get('_browser_state')
    return _decode_browser_state(raw)


def _hydrate_browser_state(estado):
    estado = estado if isinstance(estado, dict) else {}
    for key in _SESSION_KEYS:
        default = {} if key in DICT_KEYS else []
        value = estado.get(key, default)
        if key in DICT_KEYS:
            session[key] = copy.deepcopy(value if isinstance(value, dict) else default)
        else:
            session[key] = copy.deepcopy(value if isinstance(value, list) else default)
    session['tema'] = estado.get('tema', 'claro') if estado.get('tema') in {'claro', 'escuro'} else 'claro'
    resultados = estado.get(_BROWSER_RESULTS_KEY, {})
    session[_BROWSER_RESULTS_KEY] = copy.deepcopy(resultados if isinstance(resultados, dict) else {})
    session.modified = True


def _browser_snapshot():
    data = {key: copy.deepcopy(session.get(key, {} if key in DICT_KEYS else [])) for key in _SESSION_KEYS}
    data['tema'] = session.get('tema', 'claro') if session.get('tema') in {'claro', 'escuro'} else 'claro'
    resultados = session.get(_BROWSER_RESULTS_KEY, {})
    data[_BROWSER_RESULTS_KEY] = copy.deepcopy(resultados if isinstance(resultados, dict) else {})
    data['_schema_version'] = 2
    data['_saved_at'] = __import__('datetime').datetime.utcnow().isoformat(timespec='seconds') + 'Z'
    return data


def _hydrate_persisted_state():
    """Carrega o estado grande para uso apenas durante a requisição atual."""
    if STORAGE_MODE == 'browser':
        _hydrate_browser_state(_browser_state_from_request())
        return
    estado = load_state(_workspace_id())
    for key in _SESSION_KEYS:
        session[key] = copy.deepcopy(estado.get(key, {} if key in DICT_KEYS else []))
    if estado.get('tema') in {'claro', 'escuro'}:
        session['tema'] = estado['tema']
    if estado.get('resultado_token'):
        session['resultado_token'] = estado['resultado_token']
    elif 'resultado_token' in session:
        session.pop('resultado_token', None)


@app.before_request
def _seguranca_basica():
    if request.endpoint == 'static':
        return None
    _workspace_id()
    token = _csrf_token()
    try:
        _hydrate_persisted_state()
    except ValueError as exc:
        return jsonify({'status': 'erro', 'mensagem': str(exc)}), 400
    if request.method in {'POST', 'PUT', 'PATCH', 'DELETE'}:
        supplied = request.headers.get('X-CSRF-Token') or request.form.get('csrf_token', '')
        if not supplied or not hmac.compare_digest(str(supplied), str(token)):
            return jsonify({
                'status': 'erro',
                'mensagem': 'Sessão renovada automaticamente. Tentando salvar novamente...',
                'csrf_token': token,
                'csrf_refresh': True,
            }), 403
    return None


@app.after_request
def _minimizar_cookie_da_sessao(response):
    """Mantém cookies pequenos e devolve o snapshot ao IndexedDB no modo web."""
    if request.endpoint != 'static' and STORAGE_MODE == 'browser':
        # Cookie CSRF próprio: independente da chave temporária de uma função serverless.
        response.set_cookie(
            'combinix_csrf', _csrf_token(), max_age=60 * 60 * 24 * 365,
            secure=bool(request.is_secure), httponly=True, samesite='Lax'
        )
        if (request.method in {'POST', 'PUT', 'PATCH', 'DELETE'}
                and response.is_json and not getattr(response, 'direct_passthrough', False)):
            payload = response.get_json(silent=True)
            if isinstance(payload, dict):
                payload['browser_state'] = _browser_snapshot()
                payload['storage_mode'] = 'browser'
                response.set_data(json.dumps(payload, ensure_ascii=False))
                response.headers['Content-Type'] = 'application/json; charset=utf-8'
                response.headers['Content-Length'] = str(len(response.get_data()))
    for key in _SESSION_KEYS:
        session.pop(key, None)
    session.pop(_BROWSER_RESULTS_KEY, None)
    if STORAGE_MODE == 'browser':
        session.pop('resultado_token', None)
    return response


@app.context_processor
def _inject_security_context():
    browser_state_bootstrap = None
    if STORAGE_MODE == 'browser' and request.method == 'POST' and request.endpoint in {'index', 'config'}:
        browser_state_bootstrap = _browser_snapshot()
    return {
        'csrf_token': _csrf_token(),
        'app_version': APP_VERSION,
        'storage_mode': STORAGE_MODE,
        'is_browser_storage': STORAGE_MODE == 'browser',
        'browser_state_bootstrap': browser_state_bootstrap,
    }

@app.errorhandler(RequestEntityTooLarge)
def _arquivo_grande(_exc):
    return jsonify({'status':'erro','mensagem':'Arquivo muito grande. O estado ou backup JSON deve ter no máximo 16 MiB.'}), 413

# ═══════════════════════════════════════════════════════════════════════════
# ARMAZENAMENTO DE RESULTADOS — fora da session (evita cookie >4KB)
# ═══════════════════════════════════════════════════════════════════════════
# A session armazena APENAS um token UUID (poucos bytes). Os resultados grandes
# (grades, relatórios) ficam em memória do servidor e persistidos em disco.
# Isso evita o erro "session cookie too large" que causava tela em branco.
RESULTADOS_STORE = {}
_APP_DIR = os.path.dirname(os.path.abspath(__file__))
RESULTADOS_FILE = os.path.join(_APP_DIR, 'database', 'resultados.json')

def _load_resultados_from_disk():
    if STORAGE_MODE != 'local':
        return
    """
    Carrega store de resultados do disco (sobrevive reinício do servidor).
    Resiliente a JSON corrompido — em caso de erro, ignora o arquivo (não crasha).
    """
    if not os.path.exists(RESULTADOS_FILE):
        return
    try:
        with open(RESULTADOS_FILE, 'r', encoding='utf-8') as f:
            content = f.read()
        if not content.strip():
            return
        data = json.loads(content)
        if isinstance(data, dict):
            RESULTADOS_STORE.update(data)
    except (json.JSONDecodeError, OSError, UnicodeDecodeError) as e:
        logging.warning('[Combinix] resultados.json corrompido, ignorando: %s', e)
        # Renomear o arquivo corrompido para investigação posterior
        try:
            os.replace(RESULTADOS_FILE, RESULTADOS_FILE + '.corrupted')
        except OSError:
            pass

def _save_resultados_to_disk():
    if STORAGE_MODE != 'local':
        return
    """
    Persiste store em disco de forma ATÔMICA:
    escreve em arquivo temporário e usa os.replace (rename atômico).
    Isso evita 'Extra data' por escritas interrompidas.
    """
    import tempfile
    try:
        target_dir = os.path.dirname(RESULTADOS_FILE)
        os.makedirs(target_dir, exist_ok=True)
        fd, tmp_path = tempfile.mkstemp(dir=target_dir, prefix='.tmp_res_', suffix='.json')
        try:
            with os.fdopen(fd, 'w', encoding='utf-8') as f:
                json.dump(RESULTADOS_STORE, f, ensure_ascii=False)
                f.flush()
                try: os.fsync(f.fileno())
                except OSError: pass
            os.replace(tmp_path, RESULTADOS_FILE)
        except Exception:
            if os.path.exists(tmp_path):
                try: os.remove(tmp_path)
                except OSError: pass
            raise
    except Exception as e:
        logging.warning('[Combinix] Falha salvar resultados: %s', e)

def _novo_token():
    return uuid.uuid4().hex[:16]  # 16 chars (8 bytes) é suficiente

def _store_resultados(dados):
    """Armazena resultados no disco local ou no snapshot devolvido ao navegador."""
    if STORAGE_MODE == 'browser':
        session[_BROWSER_RESULTS_KEY] = copy.deepcopy(dados if isinstance(dados, dict) else {})
        session.modified = True
        return 'browser'
    token_antigo = session.get('resultado_token')
    if token_antigo and token_antigo in RESULTADOS_STORE:
        del RESULTADOS_STORE[token_antigo]
    token = _novo_token()
    RESULTADOS_STORE[token] = {'workspace_id': _workspace_id(), 'dados': dados}
    _save_resultados_to_disk()
    return token

def _get_resultados():
    """Retorna resultados do workspace local ou do IndexedDB enviado pelo navegador."""
    if STORAGE_MODE == 'browser':
        dados = session.get(_BROWSER_RESULTS_KEY, {})
        return copy.deepcopy(dados if isinstance(dados, dict) else {})
    token = session.get('resultado_token')
    if not token or token not in RESULTADOS_STORE:
        return {}
    registro = RESULTADOS_STORE[token]
    # Migração transparente de versões antigas, válidas apenas para o workspace local.
    if isinstance(registro, dict) and 'dados' not in registro:
        return registro if _workspace_id() == 'local' else {}
    if not isinstance(registro, dict) or registro.get('workspace_id') != _workspace_id():
        return {}
    dados = registro.get('dados', {})
    return dados if isinstance(dados, dict) else {}

def _limpar_resultados_usuario():
    if STORAGE_MODE == 'browser':
        session[_BROWSER_RESULTS_KEY] = {}
        session.pop('resultado_token', None)
        session.modified = True
        return
    token = session.get('resultado_token')
    if token and token in RESULTADOS_STORE:
        del RESULTADOS_STORE[token]
        _save_resultados_to_disk()
    session.pop('resultado_token', None)

_load_resultados_from_disk()

# ─── Caminhos ABSOLUTOS — funciona independente do diretório de execução ────
# Sempre relativo ao diretório onde app.py está, não ao diretório de trabalho
BASE_DIR           = os.path.dirname(os.path.abspath(__file__))
DISCIPLINAS_FOLDER = os.path.join(BASE_DIR, 'disciplinas')
PROFESSORES_FOLDER = os.path.join(BASE_DIR, 'professores')

os.makedirs(DISCIPLINAS_FOLDER, exist_ok=True)
os.makedirs(PROFESSORES_FOLDER, exist_ok=True)

def _catalog_names(folder):
    try:
        return sorted(os.path.splitext(name)[0] for name in os.listdir(folder) if name.endswith('.json'))
    except OSError:
        return []

def _safe_catalog_file(folder, catalog_name):
    """Aceita somente nomes presentes no catálogo; impede ../ e leitura arbitrária."""
    if not isinstance(catalog_name, str) or catalog_name not in _catalog_names(folder):
        return None
    base = os.path.realpath(folder)
    candidate = os.path.realpath(os.path.join(folder, catalog_name + '.json'))
    try:
        if os.path.commonpath([base, candidate]) != base:
            return None
    except ValueError:
        return None
    return candidate

def _load_catalog_list(path):
    if not path:
        return []
    with open(path, 'r', encoding='utf-8') as handle:
        data = json.load(handle)
    if not isinstance(data, list):
        raise ValueError('O arquivo de catálogo deve conter uma lista JSON.')
    return data

def _int_field(value, name, minimum, maximum):
    try:
        number = int(value)
    except (TypeError, ValueError):
        raise ValueError(f'{name} deve ser um número inteiro.')
    if number < minimum or number > maximum:
        raise ValueError(f'{name} deve ficar entre {minimum} e {maximum}.')
    return number

def _sanitize_disciplina(raw, default_curso='Geral', permitir_carga_zero=False):
    if not isinstance(raw, dict):
        raise ValueError('Disciplina inválida.')
    nome = str(raw.get('nome', '')).strip()
    if not nome:
        raise ValueError('Informe o nome da disciplina.')
    carga_minima = 0 if permitir_carga_zero else 1
    return {
        'nome': nome[:180],
        'codigo': str(raw.get('codigo', 'MAN') or 'MAN').strip()[:40],
        'curso': str(raw.get('curso', default_curso) or default_curso).strip()[:100],
        'semestre': _int_field(raw.get('semestre', 1), 'Semestre', 1, 30),
        'carga_horaria': _int_field(raw.get('carga_horaria', raw.get('carga', 60)), 'Carga horária', carga_minima, 1000),
    }

def _sanitize_professor(raw):
    nome = str((raw or {}).get('nome', '') if isinstance(raw, dict) else raw).strip()
    if not nome:
        raise ValueError('Informe o nome do professor.')
    return {'nome': nome[:180]}

def _safe_index(value, total, field='Índice'):
    try:
        idx = int(value)
    except (TypeError, ValueError):
        raise ValueError(f'{field} inválido.')
    if not 0 <= idx < total:
        raise ValueError(f'{field} inválido.')
    return idx

def _sanitize_slots_input(raw):
    if raw is None:
        return []
    if not isinstance(raw, list):
        raise ValueError('A lista de horários deve ser válida.')
    vistos, slots = set(), []
    for item in raw[:300]:
        if not isinstance(item, (list, tuple)) or len(item) != 2:
            raise ValueError('Formato de horário inválido.')
        dia, hora = str(item[0]), str(item[1])
        if dia not in DIAS or hora not in HORARIOS:
            raise ValueError(f'Horário inválido: {dia} {hora}.')
        if (dia, hora) not in vistos:
            vistos.add((dia, hora)); slots.append([dia, hora])
    return slots

def _sanitize_disc_indices(raw, total):
    if raw is None:
        return []
    if not isinstance(raw, list):
        raise ValueError('A lista de disciplinas do professor deve ser válida.')
    saida=[]
    for item in raw:
        idx=_safe_index(item,total,'Índice de disciplina')
        if idx not in saida: saida.append(idx)
    return saida

def _sanitize_prof_names(raw, allowed_names=None, strict=True):
    if raw is None:
        return []
    if not isinstance(raw, list):
        raise ValueError('A lista de professores fixos deve ser válida.')
    has_allowlist = allowed_names is not None
    allowed = set(allowed_names or [])
    saida=[]
    for item in raw:
        nome=str(item).strip()
        if not nome:
            continue
        if has_allowlist and nome not in allowed:
            if strict:
                raise ValueError(f'Professor fixo inválido: {nome}.')
            continue
        if nome not in saida:
            saida.append(nome)
    return saida

def _sincronizar_professores_fixos():
    """Mantém reservas por disciplina coerentes com a aba Professores.

    Uma disciplina com professor(es) fixo(s) pertence somente a eles. As
    configurações dos demais docentes não podem reintroduzi-la silenciosamente.
    """
    disciplinas=session.get('disciplinas_selecionadas', [])
    professores=session.get('professores_selecionados', [])
    cfg_disc=session.get('config_disciplinas', [])
    cfg_prof=session.get('config_professores', [])
    nomes=[str(p.get('nome','')).strip() for p in professores if str(p.get('nome','')).strip()]
    nomes_set=set(nomes)
    while len(cfg_disc)<len(disciplinas): cfg_disc.append(_cfg_padrao(disciplinas[len(cfg_disc)]))
    while len(cfg_prof)<len(professores): cfg_prof.append(_cfg_padrao_prof())
    reservas={}
    for di,cfg in enumerate(cfg_disc):
        if not isinstance(cfg,dict):
            cfg=_cfg_padrao(disciplinas[di] if di<len(disciplinas) else {})
            cfg_disc[di]=cfg
        fixos=_sanitize_prof_names(cfg.get('professores_fixos', []), nomes_set, strict=False)
        # Disciplinas externas pertencem a outros cursos: não reservam docentes
        # selecionados neste workspace e aparecem como “Professor externo”.
        if cfg.get('tipo', 'interna') == 'externa':
            fixos=[]
            cfg['permitir_multiplos_professores']=False
        cfg['professores_fixos']=fixos
        if len(fixos)>1:
            cfg['permitir_multiplos_professores']=True
        reservas[di]=set(fixos)
    for pi,prof in enumerate(professores):
        nome=str(prof.get('nome','')).strip()
        cfg=cfg_prof[pi] if pi<len(cfg_prof) and isinstance(cfg_prof[pi],dict) else _cfg_padrao_prof()
        cfg.setdefault('carga_alvo', cfg.get('carga_maxima',20))
        atuais=_sanitize_disc_indices(cfg.get('disciplinas_internas', []), len(disciplinas))
        filtradas=[]
        for di in atuais:
            disc_cfg=cfg_disc[di] if 0 <= di < len(cfg_disc) and isinstance(cfg_disc[di],dict) else {}
            if disc_cfg.get('tipo', 'interna') == 'externa':
                continue
            fixos=reservas.get(di,set())
            if not fixos or nome in fixos:
                filtradas.append(di)
        for di,fixos in reservas.items():
            disc_cfg=cfg_disc[di] if 0 <= di < len(cfg_disc) and isinstance(cfg_disc[di],dict) else {}
            if disc_cfg.get('tipo', 'interna') != 'externa' and nome in fixos and di not in filtradas:
                filtradas.append(di)
        cfg['disciplinas_internas']=sorted(filtradas)
        cfg_prof[pi]=cfg
    session['config_disciplinas']=cfg_disc
    session['config_professores']=cfg_prof
    session.modified=True

def _upd_cfg_prof(idx, data):
    cfgs=session.get('config_professores', [])
    idx=_safe_index(idx,len(cfgs),'Índice de professor')
    total_disc=len(session.get('disciplinas_selecionadas', []))
    maxima=_int_field(data.get('carga_maxima', 20), 'Carga máxima', 0, 100)
    alvo=_int_field(data.get('carga_alvo', maxima), 'Carga alvo', 0, 100)
    if alvo>maxima:
        raise ValueError('A carga alvo não pode ser maior do que a carga máxima.')
    cfgs[idx].update({
        'disciplinas_internas': _sanitize_disc_indices(data.get('disciplinas_internas', []), total_disc),
        'carga_alvo': alvo,
        'carga_maxima': maxima,
        'disponibilidade': _sanitize_slots_input(data.get('disponibilidade', [])),
    })
    session['config_professores']=cfgs; session.modified=True
    _sincronizar_professores_fixos()


def _sanitize_imported_disc_config(raw, disciplina, allowed_professor_names=None):
    cfg = _cfg_padrao(disciplina)
    raw = raw if isinstance(raw, dict) else {}
    cfg.update({
        'tipo': raw.get('tipo', 'interna') if raw.get('tipo', 'interna') in {'interna','externa','cedida'} else 'interna',
        'aulas_semanais': _int_field(raw.get('aulas_semanais', cfg['aulas_semanais']), 'Aulas por semana', 1, 40),
        'semestre_oferta': _int_field(raw.get('semestre_oferta', cfg['semestre_oferta']), 'Semestre de oferta', 1, 30),
        'fixacoes': _sanitize_slots_input(raw.get('fixacoes', [])),
        'restricoes': _sanitize_slots_input(raw.get('restricoes', [])),
        'permitir_multiplos_professores': bool(raw.get('permitir_multiplos_professores', False)),
        'professores_fixos': _sanitize_prof_names(raw.get('professores_fixos', []), allowed_professor_names, strict=False),
    })
    if cfg.get('tipo') == 'externa':
        cfg['professores_fixos'] = []
        cfg['permitir_multiplos_professores'] = False
    elif len(cfg.get('professores_fixos', [])) > 1:
        cfg['permitir_multiplos_professores'] = True
    # Mantém um estado único por célula. Em backups legados contraditórios, a fixação prevalece.
    fixas = {tuple(x) for x in cfg['fixacoes']}
    cfg['restricoes'] = [x for x in cfg['restricoes'] if tuple(x) not in fixas]
    if len(cfg['fixacoes']) > cfg['aulas_semanais']:
        raise ValueError(f"Backup inválido: a disciplina {disciplina.get('nome','?')} possui mais fixações do que aulas semanais.")
    return cfg


def _sanitize_imported_prof_config(raw, total_disc):
    raw = raw if isinstance(raw, dict) else {}
    maxima = _int_field(raw.get('carga_maxima', 20), 'Carga máxima', 0, 100)
    alvo = _int_field(raw.get('carga_alvo', maxima), 'Carga alvo', 0, 100)
    if alvo > maxima:
        raise ValueError('Backup inválido: carga alvo maior do que a carga máxima de um professor.')
    return {
        'disciplinas_internas': _sanitize_disc_indices(raw.get('disciplinas_internas', []), total_disc),
        'carga_alvo': alvo,
        'carga_maxima': maxima,
        'disponibilidade': _sanitize_slots_input(raw.get('disponibilidade', [])),
    }


def _sanitize_imported_groups(raw_groups, disciplinas):
    if raw_groups is None:
        return []
    if not isinstance(raw_groups, list):
        raise ValueError('Backup inválido: grupos de conflito devem formar uma lista.')
    valid_names = {str(d.get('nome','')).strip() for d in disciplinas}
    saida, nomes_vistos = [], set()
    for raw in raw_groups[:300]:
        if not isinstance(raw, dict):
            raise ValueError('Backup inválido: grupo de conflito malformado.')
        nome = str(raw.get('nome','')).strip()[:120]
        disciplinas_grupo = raw.get('disciplinas', [])
        if not isinstance(disciplinas_grupo, list):
            raise ValueError('Backup inválido: disciplinas de um grupo de conflito devem formar uma lista.')
        unicas = list(dict.fromkeys(str(x).strip() for x in disciplinas_grupo if str(x).strip() in valid_names))
        if nome and len(unicas) >= 2 and nome.casefold() not in nomes_vistos:
            nomes_vistos.add(nome.casefold())
            saida.append({'nome': nome, 'disciplinas': unicas})
    return saida

# ─── Verificação da camada de dados na inicialização ────────────────────────
def _verificar_dados():
    """
    Verifica se as pastas de dados têm arquivos JSON.
    Retorna dict com status da data layer.
    NOTA: reset() NÃO apaga esses arquivos — apenas limpa o estado do usuário.
    """
    disc_files = [f for f in os.listdir(DISCIPLINAS_FOLDER) if f.endswith('.json')]
    prof_files = [f for f in os.listdir(PROFESSORES_FOLDER) if f.endswith('.json')]
    status = {
        'disciplinas': {'pasta': DISCIPLINAS_FOLDER, 'arquivos': disc_files, 'ok': len(disc_files) > 0},
        'professores':  {'pasta': PROFESSORES_FOLDER, 'arquivos': prof_files, 'ok': len(prof_files) > 0},
    }
    if not status['disciplinas']['ok']:
        logging.warning('[Combinix] ATENÇÃO: pasta disciplinas está vazia: %s', DISCIPLINAS_FOLDER)
    if not status['professores']['ok']:
        logging.warning('[Combinix] ATENÇÃO: pasta professores está vazia: %s', PROFESSORES_FOLDER)
    return status

# Executar verificação ao iniciar o módulo
_STATUS_DADOS = _verificar_dados()

DIAS_BASE = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta']
DIAS_SABADO = DIAS_BASE + ['Sábado']
DIAS = DIAS_SABADO  # Constante completa; solver filtra conforme estado_sabado
HORARIOS = ['08:00-09:00','09:00-10:00','10:00-11:00','11:00-12:00',
            '14:00-15:00','15:00-16:00','16:00-17:00','17:00-18:00']
HORARIO_ULTIMO = '17:00-18:00'

DICT_KEYS = {'config_avancadas'}

# ─── Identidade única ──────────────────────────────────────────────────────
def disc_uid(d): return solver_disc_uid(d)
def grupo_key(c,s): return solver_grupo_key(c,s)

# ─── Config padrão ─────────────────────────────────────────────────────────
def _cfg_padrao(d):
    return solver_cfg_padrao(d)
def _cfg_padrao_prof():
    return solver_cfg_padrao_prof()
def _normalizar_avancadas(cfg):
    normalizada = solver_normalizar_avancadas(cfg)
    cfg.clear(); cfg.update(normalizada)
    return cfg

# ─── Persistência ──────────────────────────────────────────────────────────
# A session agora contém APENAS:
#   - disciplinas_selecionadas, professores_selecionados
#   - config_disciplinas, config_professores, grupos_choque, config_avancadas
#   - resultado_token (UUID curto apontando para RESULTADOS_STORE)
#   - tema
# Os resultados GRANDES ficam em RESULTADOS_STORE (em memória + disco).
_SESSION_KEYS = ['disciplinas_selecionadas', 'professores_selecionados',
                 'config_disciplinas', 'config_professores',
                 'grupos_choque', 'config_avancadas']

def auto_save():
    """Persiste o estado atual e confirma a escrita no disco.

    Antes, as rotas retornavam sucesso mesmo quando o sistema operacional não
    conseguia gravar o JSON local. Isso fazia a etapa de seleção aparentar estar
    salva e a configuração abrir vazia. Agora a falha é propagada para a
    interface com uma mensagem clara.
    """
    data = {k: session.get(k, {} if k in DICT_KEYS else []) for k in _SESSION_KEYS}
    data['tema'] = session.get('tema', 'claro')
    if STORAGE_MODE == 'browser':
        # O after_request devolve o snapshot atualizado para o IndexedDB.
        return data
    data['resultado_token'] = session.get('resultado_token', '')
    if not save_state(data, _workspace_id()):
        raise OSError('Não foi possível gravar o estado local. Verifique se a pasta do Combinix permite escrita.')
    return data

def carregar_estado_inicial():
    if STORAGE_MODE == 'browser':
        # A página foi hidratada pelo snapshot enviado pelo loader do IndexedDB.
        return any(session.get(k) for k in _SESSION_KEYS)
    estado = load_state(_workspace_id())
    if estado:
        for k in _SESSION_KEYS:
            session[k] = estado.get(k, {} if k in DICT_KEYS else [])
        session['tema'] = estado.get('tema', 'claro')
        if estado.get('resultado_token'):
            session['resultado_token'] = estado['resultado_token']
        session.modified = True
        return True
    return False

def _preservar_configs_disc(novas, antigas, cfgs_antigas):
    """CRÍTICO: Preserva configs ao navegar — não apaga dados."""
    mapa = {disc_uid(antigas[i]): cfgs_antigas[i] for i in range(min(len(antigas),len(cfgs_antigas)))}
    return [mapa.get(disc_uid(d), _cfg_padrao(d)) for d in novas]

def _preservar_configs_prof(novos, antigos, cfgs_antigas):
    mapa = {antigos[i].get('nome',''): cfgs_antigas[i] for i in range(min(len(antigos),len(cfgs_antigas)))}
    return [mapa.get(p.get('nome',''), _cfg_padrao_prof()) for p in novos]

def _remap_prof_disc_indices(disciplinas_antigas, disciplinas_novas, cfgs_prof):
    """Preserva vínculos docentes quando disciplinas são removidas ou reordenadas."""
    old_uid_by_idx = {i: disc_uid(d) for i, d in enumerate(disciplinas_antigas)}
    new_idx_by_uid = {disc_uid(d): i for i, d in enumerate(disciplinas_novas)}
    for cfg in cfgs_prof:
        if not isinstance(cfg, dict):
            continue
        novos=[]
        for raw_idx in cfg.get('disciplinas_internas', []) or []:
            try: old_idx=int(raw_idx)
            except (TypeError, ValueError): continue
            uid=old_uid_by_idx.get(old_idx)
            if uid in new_idx_by_uid and new_idx_by_uid[uid] not in novos:
                novos.append(new_idx_by_uid[uid])
        cfg['disciplinas_internas']=novos
    return cfgs_prof

# ─── Páginas ───────────────────────────────────────────────────────────────
def _browser_loader_if_needed(target, download=False):
    if STORAGE_MODE == 'browser' and request.method == 'GET':
        return render_template('browser_loader.html', target=target, download=download)
    return None


@app.route('/', methods=['GET', 'POST'])
def index():
    loader = _browser_loader_if_needed(request.full_path.rstrip('?'))
    if loader is not None:
        return loader
    if 'disciplinas_selecionadas' not in session: carregar_estado_inicial()
    return render_template('index.html')

@app.route('/config', methods=['GET', 'POST'])
def config():
    loader = _browser_loader_if_needed(request.full_path.rstrip('?'))
    if loader is not None:
        return loader
    if 'disciplinas_selecionadas' not in session: carregar_estado_inicial()
    disciplinas = session.get('disciplinas_selecionadas',[])
    professores = session.get('professores_selecionados',[])
    config_disc = session.get('config_disciplinas',[])
    config_prof = session.get('config_professores',[])
    while len(config_disc) < len(disciplinas): config_disc.append(_cfg_padrao(disciplinas[len(config_disc)]))
    while len(config_prof) < len(professores): config_prof.append(_cfg_padrao_prof())
    grupos_choque = session.get('grupos_choque',[])
    config_avancadas = session.get('config_avancadas',{'estado_sabado':'desativado','nivel_restricao':3})
    _normalizar_avancadas(config_avancadas)
    # Migra configurações antigas e sincroniza reservas docentes.
    session['config_disciplinas'] = config_disc
    session['config_professores'] = config_prof
    _sincronizar_professores_fixos()
    config_disc = session.get('config_disciplinas', [])
    config_prof = session.get('config_professores', [])
    cobertura_professores = solver_analisar_cobertura_professores(disciplinas, professores, config_disc, config_prof)
    session['config_avancadas'] = config_avancadas
    session.modified = True
    auto_save()
    return_to = request.args.get('return', '')
    if return_to not in {'/generate', '/resultados'}:
        return_to = ''
    return render_template('config.html', disciplinas=disciplinas, professores=professores,
                           config_disciplinas=config_disc, config_professores=config_prof,
                           cobertura_professores=cobertura_professores,
                           grupos_choque=grupos_choque, config_avancadas=config_avancadas,
                           return_to=return_to, selecao_restaurada=request.args.get('restaurado') == '1',
                           dias=DIAS, horarios=HORARIOS)

@app.route('/generate', methods=['GET', 'POST'])
def generate():
    loader = _browser_loader_if_needed(request.full_path.rstrip('?'))
    if loader is not None:
        return loader
    if 'disciplinas_selecionadas' not in session: carregar_estado_inicial()
    res = _get_resultados()
    return render_template('generate.html',
        show_results=bool(res.get('grade_disciplinas')),
        grade_disciplinas=res.get('grade_disciplinas',{}),
        grade_professores=res.get('grade_professores',{}),
        horarios_professores=res.get('horarios_professores',{}),
        grade_por_semestre=res.get('grade_por_semestre',{}),
        grade_por_professor=res.get('grade_por_professor',{}),
        grades_com_prof=res.get('grades_com_prof',{}),
        relatorio=res.get('relatorio',{}), alteracoes_salvas=request.args.get('alteracoes') == '1', dias=DIAS, horarios=HORARIOS)

@app.route('/resultados', methods=['GET', 'POST'])
def resultados():
    loader = _browser_loader_if_needed(request.full_path.rstrip('?'))
    if loader is not None:
        return loader
    res = _get_resultados()
    return render_template('generate.html', show_results=True,
        grade_disciplinas=res.get('grade_disciplinas',{}),
        grade_professores=res.get('grade_professores',{}),
        horarios_professores=res.get('horarios_professores',{}),
        grade_por_semestre=res.get('grade_por_semestre',{}),
        grade_por_professor=res.get('grade_por_professor',{}),
        grades_com_prof=res.get('grades_com_prof',{}),
        relatorio=res.get('relatorio',{}), alteracoes_salvas=False, dias=DIAS, horarios=HORARIOS)

# ─── APIs ──────────────────────────────────────────────────────────────────
@app.route('/api/status')
def api_status():
    """
    Endpoint de saúde: retorna status da data layer.
    O frontend usa isso para mostrar erros claros quando dados faltam.
    """
    status = _verificar_dados()  # Sempre relê do disco, não usa cache
    return jsonify({
        'ok': status['disciplinas']['ok'] and status['professores']['ok'],
        'storage_mode': STORAGE_MODE,
        'disciplinas': {
            'total': len(status['disciplinas']['arquivos']),
            'cursos': sorted(f.replace('.json','') for f in status['disciplinas']['arquivos']),
            'ok': status['disciplinas']['ok'],
        },
        'professores': {
            'total': len(status['professores']['arquivos']),
            'institutos': sorted(f.replace('.json','') for f in status['professores']['arquivos']),
            'ok': status['professores']['ok'],
        },
    })

@app.route('/api/cursos/disciplinas')
def api_cursos_disciplinas():
    try:
        cursos = _catalog_names(DISCIPLINAS_FOLDER)
    except OSError:
        cursos = []
    return jsonify(cursos)

@app.route('/api/cursos/professores')
def api_cursos_professores():
    try:
        cursos = _catalog_names(PROFESSORES_FOLDER)
    except OSError:
        cursos = []
    return jsonify(cursos)

@app.route('/api/disciplinas/<path:curso>')
def api_disciplinas(curso):
    path = _safe_catalog_file(DISCIPLINAS_FOLDER, curso)
    if not path:
        return jsonify({'erro':'Catálogo não encontrado.'}), 404
    try:
        # Componentes curriculares com carga horária zero são válidos no
        # catálogo (por exemplo, exames e TCC sem encontro semanal), mas não
        # podem ser adicionados à grade recorrente. A interface os exibe como
        # itens informativos e desativa a seleção.
        data = [_sanitize_disciplina(d, curso, permitir_carga_zero=True) for d in _load_catalog_list(path)]
        return jsonify(data)
    except (OSError, ValueError, json.JSONDecodeError) as exc:
        logging.warning('[Combinix] Catálogo de disciplinas inválido: %s', exc)
        return jsonify({'erro':'Catálogo de disciplinas inválido.', 'mensagem':str(exc)}), 422

@app.route('/api/professores/<path:curso>')
def api_professores(curso):
    path = _safe_catalog_file(PROFESSORES_FOLDER, curso)
    if not path:
        return jsonify({'erro':'Catálogo não encontrado.'}), 404
    try:
        return jsonify([_sanitize_professor(p) for p in _load_catalog_list(path)])
    except (OSError, ValueError, json.JSONDecodeError) as exc:
        logging.warning('[Combinix] Catálogo de professores inválido: %s', exc)
        return jsonify({'erro':'Catálogo de professores inválido.', 'mensagem':str(exc)}), 422

# ─── Seleção ───────────────────────────────────────────────────────────────
def _sanitize_selected_disciplines(raw):
    if not isinstance(raw, list):
        raise ValueError('A seleção de disciplinas deve ser uma lista.')
    vistos, unicas = set(), []
    for item in raw[:500]:
        disc = _sanitize_disciplina(item)
        uid = disc_uid(disc)
        if uid not in vistos:
            vistos.add(uid)
            unicas.append(disc)
    return unicas


def _sanitize_selected_professors(raw):
    if not isinstance(raw, list):
        raise ValueError('A seleção de professores deve ser uma lista.')
    vistos, professores = set(), []
    for item in raw[:500]:
        prof = _sanitize_professor(item)
        if prof['nome'] not in vistos:
            vistos.add(prof['nome'])
            professores.append(prof)
    return professores


def _apply_selections(disciplinas, professores):
    antigas_disc = session.get('disciplinas_selecionadas', [])
    antigas_prof = session.get('professores_selecionados', [])
    session['config_disciplinas'] = _preservar_configs_disc(
        disciplinas, antigas_disc, session.get('config_disciplinas', []))
    cfg_prof = _preservar_configs_prof(
        professores, antigas_prof, session.get('config_professores', []))
    session['config_professores'] = _remap_prof_disc_indices(antigas_disc, disciplinas, cfg_prof)
    session['disciplinas_selecionadas'] = disciplinas
    session['professores_selecionados'] = professores
    _sincronizar_professores_fixos()
    session.modified = True
    auto_save()


@app.route('/salvar_selecoes', methods=['POST'])
def salvar_selecoes():
    """Salva disciplinas e professores em uma única escrita atômica.

    Evita a condição de corrida da versão anterior, que disparava duas
    requisições concorrentes e podia deixar apenas metade da seleção gravada.
    A resposta só confirma sucesso depois de reler o JSON persistido.
    """
    try:
        data = request.get_json(force=True, silent=True) or {}
        disciplinas = _sanitize_selected_disciplines(data.get('disciplinas', []))
        professores = _sanitize_selected_professors(data.get('professores', []))
        _apply_selections(disciplinas, professores)
        if STORAGE_MODE == 'browser':
            # A confirmação definitiva ocorre quando o frontend grava o snapshot
            # devolvido pelo after_request no IndexedDB.
            disciplinas_confirmadas = session.get('disciplinas_selecionadas', [])
            professores_confirmados = session.get('professores_selecionados', [])
        else:
            confirmado = load_state(_workspace_id())
            disciplinas_confirmadas = confirmado.get('disciplinas_selecionadas', [])
            professores_confirmados = confirmado.get('professores_selecionados', [])
            if len(disciplinas_confirmadas) != len(disciplinas) or len(professores_confirmados) != len(professores):
                raise OSError('A gravação local não pôde ser confirmada. Tente novamente ou verifique a permissão da pasta.')
        return jsonify({
            'status': 'ok',
            'persistencia_confirmada': True,
            'disciplinas_salvas': len(disciplinas_confirmadas),
            'professores_salvos': len(professores_confirmados),
        })
    except ValueError as exc:
        return jsonify({'status':'erro', 'mensagem':str(exc)}), 400
    except OSError as exc:
        logging.warning('[Combinix] Falha ao persistir seleções: %s', exc)
        return jsonify({'status':'erro', 'mensagem':str(exc)}), 500


@app.route('/api/selecoes')
def api_selecoes():
    """Expõe uma leitura pequena do estado salvo para confirmar a navegação."""
    disciplinas = session.get('disciplinas_selecionadas', [])
    professores = session.get('professores_selecionados', [])
    return jsonify({
        'status': 'ok',
        'disciplinas': disciplinas,
        'professores': professores,
        'disciplinas_salvas': len(disciplinas),
        'professores_salvos': len(professores),
    })


@app.route('/selecionar_disciplinas', methods=['POST'])
def selecionar_disciplinas():
    try:
        raw = (request.get_json(silent=True) or {}).get('disciplinas', []) if request.is_json else json.loads(request.form.get('disciplinas', '[]'))
        unicas = _sanitize_selected_disciplines(raw)
        antigas=session.get('disciplinas_selecionadas', [])
        session['config_disciplinas'] = _preservar_configs_disc(unicas, antigas, session.get('config_disciplinas', []))
        session['config_professores'] = _remap_prof_disc_indices(antigas, unicas, session.get('config_professores', []))
        session['disciplinas_selecionadas'] = unicas
        _sincronizar_professores_fixos()
        session.modified = True; auto_save()
        return jsonify({'status':'ok'})
    except (ValueError, json.JSONDecodeError) as exc:
        return jsonify({'status':'erro','mensagem':str(exc)}), 400

@app.route('/selecionar_professores', methods=['POST'])
def selecionar_professores():
    try:
        raw = (request.get_json(silent=True) or {}).get('professores', []) if request.is_json else json.loads(request.form.get('professores', '[]'))
        profs = _sanitize_selected_professors(raw)
        session['config_professores'] = _preservar_configs_prof(profs, session.get('professores_selecionados', []), session.get('config_professores', []))
        session['professores_selecionados'] = profs
        _sincronizar_professores_fixos()
        session.modified = True; auto_save()
        return jsonify({'status':'ok'})
    except (ValueError, json.JSONDecodeError) as exc:
        return jsonify({'status':'erro','mensagem':str(exc)}), 400

@app.route('/adicionar_disciplina_manual', methods=['POST'])
def adicionar_disciplina_manual():
    try:
        d = _sanitize_disciplina({
            'nome': request.form.get('nome', ''), 'codigo': request.form.get('codigo', 'MAN'),
            'curso': request.form.get('curso', 'Manual'), 'semestre': request.form.get('semestre', 1),
            'carga_horaria': request.form.get('carga', 60),
        }, 'Manual')
        discs = session.get('disciplinas_selecionadas', [])
        if any(disc_uid(x) == disc_uid(d) for x in discs):
            return jsonify({'status':'erro','mensagem':'Esta disciplina já foi adicionada.'}), 409
        discs.append(d); session['disciplinas_selecionadas'] = discs
        cfgs = session.get('config_disciplinas', []); cfgs.append(_cfg_padrao(d))
        session['config_disciplinas'] = cfgs; session.modified = True; auto_save()
        return jsonify({'status':'ok'})
    except ValueError as exc:
        return jsonify({'status':'erro','mensagem':str(exc)}), 400

@app.route('/adicionar_professor_manual', methods=['POST'])
def adicionar_professor_manual():
    try:
        novo = _sanitize_professor({'nome':request.form.get('nome', '')})
        profs = session.get('professores_selecionados', [])
        if any(str(p.get('nome','')).strip() == novo['nome'] for p in profs):
            return jsonify({'status':'erro','mensagem':'Este professor já foi adicionado.'}), 409
        profs.append(novo)
        cfgs = session.get('config_professores', []); cfgs.append(_cfg_padrao_prof())
        session['professores_selecionados'] = profs; session['config_professores'] = cfgs
        session.modified = True; auto_save(); return jsonify({'status':'ok'})
    except ValueError as exc:
        return jsonify({'status':'erro','mensagem':str(exc)}), 400

@app.route('/remover_disciplina', methods=['POST'])
def remover_disciplina():
    try:
        d=session.get('disciplinas_selecionadas',[]); c=session.get('config_disciplinas',[]); antigas=list(d)
        idx=_safe_index(request.form.get('index',-1),len(d),'Índice de disciplina')
        d.pop(idx); (c.pop(idx) if idx<len(c) else None)
        session['disciplinas_selecionadas']=d; session['config_disciplinas']=c
        session['config_professores']=_remap_prof_disc_indices(antigas,d,session.get('config_professores',[]))
        _sincronizar_professores_fixos()
        session.modified=True; auto_save(); return jsonify({'status':'ok'})
    except ValueError as exc:
        return jsonify({'status':'erro','mensagem':str(exc)}), 400

@app.route('/remover_disciplina_config', methods=['POST'])
def remover_disciplina_config():
    """Remove disciplina diretamente na tela de configuração."""
    data=request.get_json(force=True,silent=True) or {}; idx=data.get('idx')
    if isinstance(idx,str) and idx.isdigit(): idx=int(idx)
    d=session.get('disciplinas_selecionadas',[]); c=session.get('config_disciplinas',[]); antigas=list(d)
    if isinstance(idx,int) and 0<=idx<len(d):
        d.pop(idx); (c.pop(idx) if idx<len(c) else None)
        session['disciplinas_selecionadas']=d; session['config_disciplinas']=c
        session['config_professores']=_remap_prof_disc_indices(antigas,d,session.get('config_professores',[]))
        _sincronizar_professores_fixos()
        session.modified=True; auto_save(); return jsonify({'status':'ok'})
    return jsonify({'status':'erro','mensagem':'Índice inválido'})

@app.route('/remover_professor_config', methods=['POST'])
def remover_professor_config():
    """Remove professor diretamente na tela de configuração (mesma lógica simétrica)."""
    data = request.get_json(force=True, silent=True) or {}
    idx = data.get('idx')
    if isinstance(idx, str) and idx.isdigit(): idx = int(idx)
    p = session.get('professores_selecionados', [])
    c = session.get('config_professores', [])
    if isinstance(idx, int) and 0 <= idx < len(p):
        p.pop(idx)
        if idx < len(c): c.pop(idx)
        session['professores_selecionados'] = p
        session['config_professores'] = c
        _sincronizar_professores_fixos()
        session.modified = True
        auto_save()
        return jsonify({'status': 'ok'})
    return jsonify({'status': 'erro', 'mensagem': 'Índice inválido'})

@app.route('/remover_professor', methods=['POST'])
def remover_professor():
    try:
        p=session.get('professores_selecionados',[]); c=session.get('config_professores',[])
        idx=_safe_index(request.form.get('index',-1),len(p),'Índice de professor')
        p.pop(idx); (c.pop(idx) if idx<len(c) else None)
        session['professores_selecionados']=p; session['config_professores']=c
        _sincronizar_professores_fixos()
        session.modified=True; auto_save(); return jsonify({'status':'ok'})
    except ValueError as exc:
        return jsonify({'status':'erro','mensagem':str(exc)}), 400

# ─── Configuração ──────────────────────────────────────────────────────────
def _upd_cfg_disc(idx, data):
    cfgs = session.get('config_disciplinas', [])
    if not isinstance(idx, int) or not (0 <= idx < len(cfgs)):
        raise ValueError('Índice de disciplina inválido.')
    atuais = cfgs[idx] if isinstance(cfgs[idx], dict) else {}
    nomes={str(p.get('nome','')).strip() for p in session.get('professores_selecionados', []) if str(p.get('nome','')).strip()}
    tipo=data.get('tipo', 'interna') if data.get('tipo', 'interna') in {'interna','externa','cedida'} else 'interna'
    fixos=_sanitize_prof_names(data.get('professores_fixos', atuais.get('professores_fixos', [])), nomes, strict=True)
    permitir=bool(data.get('permitir_multiplos_professores', atuais.get('permitir_multiplos_professores', False)))
    if tipo == 'externa':
        fixos=[]
        permitir=False
    elif len(fixos)>1:
        permitir=True
    cfgs[idx].update({
        'tipo': tipo,
        'aulas_semanais': _int_field(data.get('aulas_semanais', 2), 'Aulas por semana', 1, 40),
        'semestre_oferta': _int_field(data.get('semestre_oferta', 1), 'Semestre de oferta', 1, 30),
        'permitir_multiplos_professores': permitir,
        'professores_fixos': fixos,
    })
    session['config_disciplinas'] = cfgs; session.modified = True
    _sincronizar_professores_fixos()

@app.route('/salvar_config_disciplina', methods=['POST'])
def salvar_config_disciplina():
    try:
        data=request.get_json(force=True,silent=True) or {}; _upd_cfg_disc(data.get('idx'),data); auto_save(); return jsonify({'status':'ok'})
    except ValueError as exc:
        return jsonify({'status':'erro','mensagem':str(exc)}), 400

@app.route('/salvar_todas_disciplinas', methods=['POST'])
def salvar_todas_disciplinas():
    try:
        data=request.get_json(force=True,silent=True) or {}
        configs=data.get('configs',[])
        if not isinstance(configs,list): raise ValueError('Configurações de disciplinas inválidas.')
        for item in configs:
            if not isinstance(item,dict): raise ValueError('Configuração de disciplina inválida.')
            _upd_cfg_disc(item.get('idx'),item)
        auto_save(); return jsonify({'status':'ok','total':len(configs)})
    except ValueError as exc:
        return jsonify({'status':'erro','mensagem':str(exc)}), 400

@app.route('/salvar_fixacao', methods=['POST'])
def salvar_fixacao():
    try:
        data=request.get_json(force=True,silent=True) or {}
        cfgs=session.get('config_disciplinas',[])
        idx=_safe_index(data.get('idx'),len(cfgs),'Índice de disciplina')
        dia, hora, tipo = str(data.get('dia','')), str(data.get('hora','')), str(data.get('tipo',''))
        if dia not in DIAS or hora not in HORARIOS:
            raise ValueError('Dia ou horário inválido.')
        if tipo not in {'fixar','restringir','limpar'}:
            raise ValueError('Ação inválida para o horário.')
        cfg=cfgs[idx]; slot=[dia,hora]
        N=_int_field(cfg.get('aulas_semanais',2),'Aulas por semana',1,40)
        nf=[s for s in _sanitize_slots_input(cfg.get('fixacoes',[])) if s!=slot]
        nr=[s for s in _sanitize_slots_input(cfg.get('restricoes',[])) if s!=slot]
        if tipo=='fixar': nf.append(slot)
        elif tipo=='restringir': nr.append(slot)
        if len(nf)>N:
            return jsonify({'status':'erro','codigo':'fixacoes_excedem_carga','mensagem':f'Esta disciplina permite no máximo {N} fixação(ões).'}), 400
        av=session.get('config_avancadas',{}); _normalizar_avancadas(av)
        if tipo == 'fixar' and dia == 'Sábado' and av.get('estado_sabado','desativado') == 'desativado':
            return jsonify({'status':'erro','codigo':'sabado_desativado','mensagem':'Ative o sábado nas Configurações Avançadas antes de fixar uma aula nesse dia.'}), 400
        dias_v=solver_get_dias_para_nivel(3, av.get('estado_sabado','desativado'))
        if len(dias_v)*len(HORARIOS)-len(nr)<N:
            return jsonify({'status':'erro','codigo':'poucos_disponiveis','mensagem':'Restrições excessivas: libere mais horários para esta disciplina.'}), 400
        cfg['fixacoes']=nf; cfg['restricoes']=nr; session['config_disciplinas']=cfgs
        session.modified=True; auto_save(); return jsonify({'status':'ok'})
    except ValueError as exc:
        return jsonify({'status':'erro','mensagem':str(exc)}), 400

@app.route('/resetar_disciplina', methods=['POST'])
def resetar_disciplina():
    try:
        data=request.get_json(force=True,silent=True) or {}
        d=session.get('disciplinas_selecionadas',[]); c=session.get('config_disciplinas',[])
        idx=_safe_index(data.get('idx'),len(c),'Índice de disciplina')
        c[idx]=_cfg_padrao(d[idx] if idx<len(d) else {}); session['config_disciplinas']=c; session.modified=True; auto_save()
        return jsonify({'status':'ok'})
    except ValueError as exc:
        return jsonify({'status':'erro','mensagem':str(exc)}), 400

@app.route('/resetar_professor', methods=['POST'])
def resetar_professor():
    try:
        data=request.get_json(force=True,silent=True) or {}; c=session.get('config_professores',[])
        idx=_safe_index(data.get('idx'),len(c),'Índice de professor')
        c[idx]=_cfg_padrao_prof(); session['config_professores']=c; _sincronizar_professores_fixos(); session.modified=True; auto_save()
        return jsonify({'status':'ok'})
    except ValueError as exc:
        return jsonify({'status':'erro','mensagem':str(exc)}), 400

@app.route('/adicionar_grupo_choque', methods=['POST'])
def adicionar_grupo_choque():
    data=request.get_json(force=True,silent=True) or {}
    nome=str(data.get('nome','') or '').strip()[:120]
    discs=data.get('disciplinas',[])
    if not isinstance(discs,list):
        return jsonify({'status':'erro','mensagem':'Lista de disciplinas inválida.'}), 400
    valid_names={str(d.get('nome','')) for d in session.get('disciplinas_selecionadas',[])}
    discs=[] if not discs else list(dict.fromkeys(str(d).strip() for d in discs if str(d).strip() in valid_names))
    if not nome or len(discs)<2:
        return jsonify({'status':'erro','mensagem':'Informe um nome e selecione ao menos duas disciplinas válidas.'}), 400
    grupos=session.get('grupos_choque',[])
    if any(str(g.get('nome','')).strip().lower()==nome.lower() for g in grupos):
        return jsonify({'status':'erro','codigo':'nome_duplicado','mensagem':'Já existe um grupo com esse nome.'}), 409
    grupos.append({'nome':nome,'disciplinas':discs}); session['grupos_choque']=grupos
    session.modified=True; auto_save(); return jsonify({'status':'ok'})

@app.route('/remover_grupo_choque', methods=['POST'])
def remover_grupo_choque():
    try:
        data=request.get_json(force=True,silent=True) or {}; grupos=session.get('grupos_choque',[])
        idx=_safe_index(data.get('idx',-1),len(grupos),'Índice de grupo')
        grupos.pop(idx); session['grupos_choque']=grupos; session.modified=True; auto_save(); return jsonify({'status':'ok'})
    except ValueError as exc:
        return jsonify({'status':'erro','mensagem':str(exc)}), 400

@app.route('/limpar_todos_conflitos', methods=['POST'])
def limpar_todos_conflitos():
    session['grupos_choque']=[]; session.modified=True; auto_save(); return jsonify({'status':'ok'})

@app.route('/salvar_config_professor', methods=['POST'])
def salvar_config_professor():
    try:
        data=request.get_json(force=True,silent=True) or {}
        _upd_cfg_prof(data.get('idx'),data); auto_save(); return jsonify({'status':'ok'})
    except ValueError as exc:
        return jsonify({'status':'erro','mensagem':str(exc)}), 400

@app.route('/salvar_todas_professores', methods=['POST'])
def salvar_todas_professores():
    try:
        data=request.get_json(force=True,silent=True) or {}
        configs=data.get('configs',[])
        if not isinstance(configs,list): raise ValueError('Configurações de professores inválidas.')
        for item in configs:
            if not isinstance(item,dict): raise ValueError('Configuração de professor inválida.')
            _upd_cfg_prof(item.get('idx'),item)
        auto_save(); return jsonify({'status':'ok','total':len(configs)})
    except ValueError as exc:
        return jsonify({'status':'erro','mensagem':str(exc)}), 400

@app.route('/salvar_config_avancadas', methods=['POST'])
def salvar_config_avancadas():
    data = request.get_json(force=True, silent=True) or {}
    nv = data.get('nivel_restricao', 3)
    try: nv = int(nv)
    except (TypeError, ValueError): nv = 3
    if nv not in (1, 2, 3): nv = 3
    cfg = {
        'estado_sabado':   data.get('estado_sabado', 'desativado'),
        'nivel_restricao': nv,
    }
    _normalizar_avancadas(cfg)
    if cfg.get('estado_sabado') == 'desativado':
        for disc_cfg in session.get('config_disciplinas', []):
            if any(isinstance(slot, (list, tuple)) and len(slot) == 2 and slot[0] == 'Sábado' for slot in disc_cfg.get('fixacoes', [])):
                return jsonify({'status':'erro','mensagem':'Não é possível desativar o sábado enquanto houver aulas fixadas nesse dia. Remova as fixações ou mantenha o sábado ativo.'}), 400
    session['config_avancadas'] = cfg
    session.modified = True; auto_save()
    return jsonify({'status': 'ok'})

@app.route('/salvar_contexto_recomendacao', methods=['POST'])
def salvar_contexto_recomendacao():
    """Salva todas as alterações visíveis antes de voltar ao diagnóstico.

    Os atalhos do relatório conduzem ao ponto exato da configuração. Este
    endpoint garante que o retorno à geração considere as alterações feitas em
    disciplinas, professores e parâmetros avançados em uma única confirmação.
    """
    snapshot={k:copy.deepcopy(session.get(k, {} if k in DICT_KEYS else [])) for k in _SESSION_KEYS}
    try:
        data=request.get_json(force=True,silent=True) or {}
        cfg_disc=data.get('config_disciplinas', [])
        cfg_prof=data.get('config_professores', [])
        avancadas=data.get('config_avancadas', {})
        if not isinstance(cfg_disc,list) or not isinstance(cfg_prof,list) or not isinstance(avancadas,dict):
            raise ValueError('Configurações inválidas.')
        for item in cfg_disc:
            if not isinstance(item,dict): raise ValueError('Configuração de disciplina inválida.')
            _upd_cfg_disc(item.get('idx'), item)
        for item in cfg_prof:
            if not isinstance(item,dict): raise ValueError('Configuração de professor inválida.')
            _upd_cfg_prof(item.get('idx'), item)
        nv=avancadas.get('nivel_restricao',3)
        try: nv=int(nv)
        except (TypeError,ValueError): nv=3
        if nv not in (1,2,3): nv=3
        cfg={'estado_sabado':avancadas.get('estado_sabado','desativado'),'nivel_restricao':nv}
        _normalizar_avancadas(cfg)
        if cfg.get('estado_sabado') == 'desativado':
            for disc_cfg in session.get('config_disciplinas', []):
                if any(isinstance(slot,(list,tuple)) and len(slot)==2 and slot[0]=='Sábado' for slot in disc_cfg.get('fixacoes', [])):
                    raise ValueError('Não é possível desativar o sábado enquanto houver aulas fixadas nesse dia.')
        session['config_avancadas']=cfg
        session.modified=True
        auto_save()
        return jsonify({'status':'ok','mensagem':'Alterações salvas. A próxima geração usará as novas regras.'})
    except ValueError as exc:
        for k,v in snapshot.items(): session[k]=v
        session.modified=True
        return jsonify({'status':'erro','mensagem':str(exc)}),400

# =============================================================================
# MOTOR DE HORÁRIOS V2 — implementação isolada em services/solver.py
# =============================================================================
def gerar_grade(variant_seed=0):
    return solve_schedule(
        session.get('disciplinas_selecionadas', []),
        session.get('professores_selecionados', []),
        session.get('config_disciplinas', []),
        session.get('config_professores', []),
        session.get('grupos_choque', []),
        session.get('config_avancadas', {}),
        variant_seed=variant_seed,
    )

def _qualidade_relatorio(relatorio):
    """Compara resultados sem trocar uma grade boa por uma alternativa inferior.

    O score heurístico pode variar entre duas grades completas igualmente úteis.
    Para regenerar, a prioridade é preservar o nível de conclusão: status,
    disciplinas alocadas e aulas alocadas. O score serve apenas para escolher a
    melhor tentativa dentro do mesmo nível de conclusão.
    """
    relatorio = relatorio if isinstance(relatorio, dict) else {}
    status=relatorio.get('status_geracao','impossivel')
    return ({'sucesso':2,'parcial':1,'impossivel':0}.get(status,0),
            int(relatorio.get('disciplinas_alocadas',0) or 0),
            int(relatorio.get('aulas_alocadas',0) or 0),
            float(relatorio.get('score',0) or 0))


def _qualidade_resultado(resultado):
    rel=resultado[5] if len(resultado)>5 and isinstance(resultado[5],dict) else {}
    return _qualidade_relatorio(rel)


def _resultado_armazenado_para_tupla(anterior):
    """Reconstrói o formato interno do solver a partir do resultado salvo."""
    if not isinstance(anterior, dict):
        return None
    rel=anterior.get('relatorio', {})
    if not isinstance(rel, dict) or not rel.get('assinatura_grade'):
        return None
    return (
        anterior.get('grade_disciplinas', {}), anterior.get('grade_professores', {}),
        anterior.get('horarios_professores', {}), anterior.get('grade_por_semestre', {}),
        anterior.get('grade_por_professor', {}), copy.deepcopy(rel), anterior.get('grades_com_prof', {}),
    )

@app.route('/iniciar_geracao', methods=['POST'])
def iniciar_geracao():
    try:
        discs = session.get('disciplinas_selecionadas', [])
        if not discs:
            return jsonify({'status':'erro',
                            'mensagem':'Nenhuma disciplina selecionada',
                            'erros':['Volte à tela de Seleção e escolha pelo menos uma disciplina.']})

        # Validação: configs existem para cada disciplina
        config_disc = session.get('config_disciplinas', [])
        if len(config_disc) < len(discs):
            session['config_disciplinas'] = [_cfg_padrao(d) for d in discs]
            session.modified = True

        payload=request.get_json(silent=True) or {}
        regenerar=bool(payload.get('regenerar'))
        anterior=_get_resultados() if regenerar else {}
        rel_anterior=anterior.get('relatorio', {}) if isinstance(anterior,dict) else {}
        assinatura_anterior=rel_anterior.get('assinatura_grade', '') if isinstance(rel_anterior,dict) else ''
        cursor_anterior=int(rel_anterior.get('cursor_regeneracao', rel_anterior.get('semente_variacao', 0)) or 0)
        semente_base=cursor_anterior + 1 if regenerar else 0
        qualidade_anterior=_qualidade_relatorio(rel_anterior)
        resultado_anterior=_resultado_armazenado_para_tupla(anterior) if regenerar else None

        tentativas=[]
        max_tentativas=5 if regenerar and assinatura_anterior else 1
        escolhido=None
        alternativa_encontrada=False
        for tentativa in range(max_tentativas):
            resultado=gerar_grade(semente_base + tentativa)
            tentativas.append(resultado)
            assinatura=(resultado[5] or {}).get('assinatura_grade','')
            qualidade=_qualidade_resultado(resultado)
            if escolhido is None or qualidade > _qualidade_resultado(escolhido):
                escolhido=resultado
            # Uma alternativa só substitui a grade visível quando mantém pelo
            # menos o mesmo nível de conclusão da anterior. O score pode mudar.
            if (regenerar and assinatura_anterior and assinatura and assinatura != assinatura_anterior
                    and qualidade[:3] >= qualidade_anterior[:3]
                    and resultado[5].get('status_geracao') != 'impossivel'):
                escolhido=resultado
                alternativa_encontrada=True
                break
        # Se nenhuma alternativa equivalente apareceu, mantém a grade anterior
        # em vez de trocar silenciosamente por uma tentativa pior ou idêntica.
        if regenerar and assinatura_anterior and not alternativa_encontrada and resultado_anterior:
            escolhido=resultado_anterior
        gd, gp, hp, gs, gpp, rel, gcp = escolhido
        if regenerar:
            ultimo_seed=semente_base + max(0, len(tentativas)-1)
            rel['cursor_regeneracao']=ultimo_seed
            if alternativa_encontrada:
                mensagem_reg='Uma combinação diferente, com o mesmo nível de conclusão da anterior, foi encontrada e exibida.'
                resultado_reg='alternativa_encontrada'
            elif assinatura_anterior:
                mensagem_reg=(f'Após {len(tentativas)} tentativa(s) adicionais, o motor não encontrou outra combinação diferente com qualidade equivalente. '
                              'A grade anterior foi mantida. Isso pode indicar uma grade muito restrita ou uma solução praticamente única, mas não é uma prova matemática de unicidade. '
                              'Ao clicar em Regerar novamente, o motor explorará novas tentativas.')
                resultado_reg='mesma_combinacao'
            else:
                mensagem_reg='A grade anterior era de uma versão antiga. Uma nova tentativa foi executada com o motor atualizado.'
                resultado_reg='nova_tentativa'
            rel['regeneracao']={'solicitada':True,'tentativas':len(tentativas),'resultado':resultado_reg,'mensagem':mensagem_reg}
        else:
            rel['cursor_regeneracao']=int(rel.get('semente_variacao', 0) or 0)
            rel['regeneracao']={'solicitada':False,'tentativas':1,'resultado':'primeira_geracao','mensagem':'Primeira combinação gerada.'}

        status_motor = rel.get('status_geracao', 'impossivel')
        if status_motor == 'impossivel':
            return jsonify({
                'status':'erro',
                'mensagem':'Não foi possível montar uma grade válida com as configurações atuais.',
                'erros':rel.get('erros') or [d.get('detalhes','') for d in rel.get('diagnosticos',[]) if d.get('severidade') == 'erro'],
                'diagnosticos':rel.get('diagnosticos', []),
                'relatorio':rel,
            }), 422

        # Armazenar no STORE (não na session) e guardar apenas o token
        token = _store_resultados({
            'grade_disciplinas':     gd,
            'grade_professores':     gp,
            'horarios_professores':  hp,
            'grade_por_semestre':    gs,
            'grade_por_professor':   gpp,
            'grades_com_prof':       gcp,
            'relatorio':             rel,
        })
        session['resultado_token'] = token
        session.modified = True
        auto_save()
        return jsonify({'status':('parcial' if rel.get('status_geracao') == 'parcial' else 'sucesso'),'relatorio':rel, 'token':token})
    except Exception as e:
        import traceback
        logging.exception('[Combinix] Falha inesperada na geração'); return jsonify({'status':'erro','mensagem':'Erro interno ao gerar a grade. Revise os dados e tente novamente.'}), 500

# ─── Download / Export / Import / Reset / Tema ────────────────────────────
@app.route('/download/<tipo>', methods=['GET', 'POST'])
def download(tipo):
    loader = _browser_loader_if_needed(request.full_path.rstrip('?'), download=True)
    if loader is not None:
        return loader
    mapa = {'disciplinas':   ('grade_disciplinas',    'grade_disciplinas.json'),
            'professores':   ('grade_professores',    'grade_professores.json'),
            'horarios':      ('horarios_professores', 'horarios_professores.json'),
            'semestre':      ('grade_por_semestre',   'grade_por_semestre.json'),
            'por_professor': ('grade_por_professor',  'grade_por_professor.json')}
    if tipo not in mapa: return jsonify({'erro':'tipo inválido'})
    chave, fn = mapa[tipo]
    res = _get_resultados()
    buf = io.BytesIO(json.dumps(res.get(chave, {}), ensure_ascii=False, indent=2).encode('utf-8'))
    buf.seek(0)
    return send_file(buf, mimetype='application/json', as_attachment=True, download_name=fn)


@app.route('/download_excel', methods=['GET', 'POST'])
def download_excel():
    loader = _browser_loader_if_needed(request.full_path.rstrip('?'), download=True)
    if loader is not None:
        return loader
    """
    Gera arquivo .xlsx com 4 abas, cada uma com uma versão da grade.
    Mantém o mesmo formato visual da interface (linhas=horários, colunas=dias).
    """
    res = _get_resultados()
    if not res.get('grade_disciplinas'):
        return jsonify({'erro': 'Nenhuma grade gerada'}), 404

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return jsonify({'erro':'Módulo openpyxl não instalado. Execute: pip install openpyxl'}), 500

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove sheet padrão

    # Estilos
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='6C9EBF')
    title_font  = Font(bold=True, size=14, color='2D2D5F')
    subtitle_font = Font(italic=True, size=10, color='666666')
    cell_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(border_style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    almoco_fill = PatternFill('solid', fgColor='F0F0F0')
    empty_fill  = PatternFill('solid', fgColor='FAFAFA')

    HORARIOS_TODOS = HORARIOS  # mesmos horários do solver

    def _draw_grid_sheet(ws, titulo, subtitulo, dias_grid, get_cell_value):
        """Desenha um cabeçalho + tabela linhas=horários, colunas=dias."""
        ws.cell(row=1, column=1, value=titulo).font = title_font
        if subtitulo:
            ws.cell(row=2, column=1, value=subtitulo).font = subtitle_font
        # Header (linha 4)
        ws.cell(row=4, column=1, value='Horário').font = header_font
        ws.cell(row=4, column=1).fill = header_fill
        ws.cell(row=4, column=1).alignment = cell_align
        ws.cell(row=4, column=1).border = border
        for j, dia in enumerate(dias_grid, start=2):
            c = ws.cell(row=4, column=j, value=dia)
            c.font = header_font; c.fill = header_fill; c.alignment = cell_align; c.border = border
        # Linhas de dados — TODOS os horários + linha de almoço
        row_idx = 5
        for i, hora in enumerate(HORARIOS_TODOS):
            # Inserir linha de almoço entre 11-12 e 14-15
            if i == 4:
                ws.cell(row=row_idx, column=1, value='12:00-14:00').font = Font(italic=True, color='888888')
                ws.cell(row=row_idx, column=1).alignment = cell_align
                ws.cell(row=row_idx, column=1).border = border
                for j in range(2, len(dias_grid)+2):
                    c = ws.cell(row=row_idx, column=j, value='almoço')
                    c.alignment = cell_align; c.fill = almoco_fill; c.border = border
                    c.font = Font(italic=True, color='888888', size=9)
                row_idx += 1
            ws.cell(row=row_idx, column=1, value=hora).font = Font(bold=True, color='555555')
            ws.cell(row=row_idx, column=1).alignment = cell_align
            ws.cell(row=row_idx, column=1).border = border
            for j, dia in enumerate(dias_grid, start=2):
                val = get_cell_value(dia, hora)
                c = ws.cell(row=row_idx, column=j, value=val if val and val != '—' else '')
                c.alignment = cell_align; c.border = border
                if not val or val == '—':
                    c.fill = empty_fill
            row_idx += 1
        # Larguras
        ws.column_dimensions['A'].width = 14
        for j in range(2, len(dias_grid)+2):
            ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = 22
        ws.row_dimensions[1].height = 24
        for r in range(4, row_idx):
            ws.row_dimensions[r].height = 28

    # ── Aba 1: Por Semestre (uma sub-tabela por grupo) ─────────────────────
    gs = res.get('grade_por_semestre', {})
    ws1 = wb.create_sheet('Por Semestre')
    if gs:
        # Para cada grupo, criamos um bloco vertical
        row_offset = 0
        for gk in sorted(gs.keys()):
            grade_g = gs[gk]
            curso, sem = (gk.split('|') + ['Geral','1'])[:2]
            titulo = f'{sem}º SEMESTRE — {curso}'
            subt = ''
            dias_grid = list(grade_g.keys())
            # Reescreve esta seção começando em row_offset+1
            base_row = row_offset + 1
            ws1.cell(row=base_row, column=1, value=titulo).font = title_font
            ws1.cell(row=base_row+2, column=1, value='Horário').font = header_font
            ws1.cell(row=base_row+2, column=1).fill = header_fill
            ws1.cell(row=base_row+2, column=1).alignment = cell_align
            ws1.cell(row=base_row+2, column=1).border = border
            for j, dia in enumerate(dias_grid, start=2):
                c = ws1.cell(row=base_row+2, column=j, value=dia)
                c.font = header_font; c.fill = header_fill; c.alignment = cell_align; c.border = border
            row = base_row + 3
            for i, hora in enumerate(HORARIOS_TODOS):
                if i == 4:
                    ws1.cell(row=row, column=1, value='12:00-14:00').alignment = cell_align
                    ws1.cell(row=row, column=1).border = border
                    for j in range(2, len(dias_grid)+2):
                        c = ws1.cell(row=row, column=j, value='almoço')
                        c.alignment = cell_align; c.fill = almoco_fill; c.border = border
                    row += 1
                ws1.cell(row=row, column=1, value=hora).alignment = cell_align
                ws1.cell(row=row, column=1).border = border
                for j, dia in enumerate(dias_grid, start=2):
                    val = grade_g.get(dia, {}).get(hora, '—')
                    c = ws1.cell(row=row, column=j, value=val if val != '—' else '')
                    c.alignment = cell_align; c.border = border
                    if val == '—': c.fill = empty_fill
                row += 1
            row_offset = row + 2  # espaço entre tabelas
        ws1.column_dimensions['A'].width = 14
        for col_letter in 'BCDEFG':
            ws1.column_dimensions[col_letter].width = 22

    # ── Aba 2: Combinada ────────────────────────────────────────────────────
    gd = res.get('grade_disciplinas', {})
    ws2 = wb.create_sheet('Grade Combinada')
    if gd:
        dias_g = list(gd.keys())
        _draw_grid_sheet(ws2, 'Grade Combinada', 'Todos os semestres em uma única tabela',
                         dias_g, lambda d, h: gd.get(d, {}).get(h, '—'))

    # ── Aba 3: Com Professores (por grupo, igual à aba 1 mas com prof) ─────
    gcp = res.get('grades_com_prof', {})
    ws3 = wb.create_sheet('Com Professores')
    if gcp:
        row_offset = 0
        for gk in sorted(gcp.keys()):
            grade_g = gcp[gk]
            curso, sem = (gk.split('|') + ['Geral','1'])[:2]
            titulo = f'{sem}º SEMESTRE — {curso} (com professores)'
            dias_grid = list(grade_g.keys())
            base_row = row_offset + 1
            ws3.cell(row=base_row, column=1, value=titulo).font = title_font
            ws3.cell(row=base_row+2, column=1, value='Horário').font = header_font
            ws3.cell(row=base_row+2, column=1).fill = header_fill
            ws3.cell(row=base_row+2, column=1).alignment = cell_align
            ws3.cell(row=base_row+2, column=1).border = border
            for j, dia in enumerate(dias_grid, start=2):
                c = ws3.cell(row=base_row+2, column=j, value=dia)
                c.font = header_font; c.fill = header_fill; c.alignment = cell_align; c.border = border
            row = base_row + 3
            for i, hora in enumerate(HORARIOS_TODOS):
                if i == 4:
                    ws3.cell(row=row, column=1, value='12:00-14:00').alignment = cell_align
                    ws3.cell(row=row, column=1).border = border
                    for j in range(2, len(dias_grid)+2):
                        c = ws3.cell(row=row, column=j, value='almoço')
                        c.alignment = cell_align; c.fill = almoco_fill; c.border = border
                    row += 1
                ws3.cell(row=row, column=1, value=hora).alignment = cell_align
                ws3.cell(row=row, column=1).border = border
                for j, dia in enumerate(dias_grid, start=2):
                    val = grade_g.get(dia, {}).get(hora, '—')
                    c = ws3.cell(row=row, column=j, value=val if val and val != '—' else '')
                    c.alignment = cell_align; c.border = border
                    if not val or val == '—': c.fill = empty_fill
                row += 1
            row_offset = row + 2
        ws3.column_dimensions['A'].width = 14
        for col_letter in 'BCDEFG':
            ws3.column_dimensions[col_letter].width = 22

    # ── Aba 4: Por Professor ────────────────────────────────────────────────
    gpp = res.get('grade_por_professor', {})
    cargas = res.get('relatorio', {}).get('carga_por_professor', {})
    ws4 = wb.create_sheet('Por Professor')
    if gpp:
        row_offset = 0
        for prof in gpp.keys():
            grade_p = gpp[prof]
            carga = cargas.get(prof, {})
            definida = carga.get('definida', '?')
            alocada  = carga.get('alocada', 0)
            dias_grid = list(grade_p.keys())
            base_row = row_offset + 1
            ws4.cell(row=base_row, column=1, value=prof).font = title_font
            ws4.cell(row=base_row+1, column=1, value=f'Carga Horária Definida: {definida}h/sem · Carga Alocada: {alocada}h/sem').font = subtitle_font
            ws4.cell(row=base_row+3, column=1, value='Horário').font = header_font
            ws4.cell(row=base_row+3, column=1).fill = header_fill
            ws4.cell(row=base_row+3, column=1).alignment = cell_align
            ws4.cell(row=base_row+3, column=1).border = border
            for j, dia in enumerate(dias_grid, start=2):
                c = ws4.cell(row=base_row+3, column=j, value=dia)
                c.font = header_font; c.fill = header_fill; c.alignment = cell_align; c.border = border
            row = base_row + 4
            for i, hora in enumerate(HORARIOS_TODOS):
                if i == 4:
                    ws4.cell(row=row, column=1, value='12:00-14:00').alignment = cell_align
                    ws4.cell(row=row, column=1).border = border
                    for j in range(2, len(dias_grid)+2):
                        c = ws4.cell(row=row, column=j, value='almoço')
                        c.alignment = cell_align; c.fill = almoco_fill; c.border = border
                    row += 1
                ws4.cell(row=row, column=1, value=hora).alignment = cell_align
                ws4.cell(row=row, column=1).border = border
                for j, dia in enumerate(dias_grid, start=2):
                    val = grade_p.get(dia, {}).get(hora, '—')
                    c = ws4.cell(row=row, column=j, value=val if val and val != '—' else '')
                    c.alignment = cell_align; c.border = border
                    if not val or val == '—': c.fill = empty_fill
                row += 1
            row_offset = row + 2
        ws4.column_dimensions['A'].width = 14
        for col_letter in 'BCDEFG':
            ws4.column_dimensions[col_letter].width = 22

    # Salvar em buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name='combinix_grade.xlsx')

@app.route('/export', methods=['GET', 'POST'])
def export():
    loader = _browser_loader_if_needed(request.full_path.rstrip('?'), download=True)
    if loader is not None:
        return loader
    # Dados leves vêm da sessão
    data = {k: session.get(k, {} if k in DICT_KEYS else []) for k in _SESSION_KEYS}
    # Resultados vêm do store
    res = _get_resultados()
    data.update({
        'resultado_grade_disciplinas':    res.get('grade_disciplinas', {}),
        'resultado_grade_professores':    res.get('grade_professores', {}),
        'resultado_horarios_professores': res.get('horarios_professores', {}),
        'resultado_grade_por_semestre':   res.get('grade_por_semestre', {}),
        'resultado_grade_por_professor':  res.get('grade_por_professor', {}),
        'resultado_grades_com_prof':       res.get('grades_com_prof', {}),
        'resultado_relatorio':            res.get('relatorio', {}),
    })
    buf = io.BytesIO(json.dumps(data, ensure_ascii=False, indent=2).encode('utf-8'))
    buf.seek(0)
    return send_file(buf, mimetype='application/json', as_attachment=True, download_name='combinix_export.json')

@app.route('/import', methods=['POST'])
def import_state():
    try:
        f = request.files.get('file')
        if not f: return jsonify({'status':'erro','mensagem':'Nenhum arquivo enviado'})
        raw = f.read()
        if len(raw) > app.config['MAX_CONTENT_LENGTH']:
            limite_mib = app.config['MAX_CONTENT_LENGTH'] // (1024 * 1024)
            return jsonify({'status':'erro','mensagem':f'Arquivo muito grande. Limite neste modo: {limite_mib} MiB.'}), 413
        data = json.loads(raw.decode('utf-8'))
        if not isinstance(data, dict):
            return jsonify({'status':'erro','mensagem':'Backup inválido: o JSON principal deve ser um objeto.'}), 400
        raw_discs = data.get('disciplinas_selecionadas', [])
        if not isinstance(raw_discs, list):
            raise ValueError('Backup inválido: disciplinas selecionadas devem formar uma lista.')
        vistos_disc, discs = set(), []
        for x in raw_discs[:500]:
            disc = _sanitize_disciplina(x)
            uid = disc_uid(disc)
            if uid not in vistos_disc:
                vistos_disc.add(uid); discs.append(disc)
        data['disciplinas_selecionadas'] = discs

        raw_profs = data.get('professores_selecionados', [])
        if not isinstance(raw_profs, list):
            raise ValueError('Backup inválido: professores selecionados devem formar uma lista.')
        vistos = set(); profs = []
        for x in raw_profs[:500]:
            prof = _sanitize_professor(x)
            if prof['nome'] not in vistos:
                vistos.add(prof['nome']); profs.append(prof)
        data['professores_selecionados'] = profs

        raw_cfg_disc = data.get('config_disciplinas', [])
        if not isinstance(raw_cfg_disc, list): raise ValueError('Backup inválido: configurações de disciplinas devem formar uma lista.')
        data['config_disciplinas'] = [_sanitize_imported_disc_config(raw_cfg_disc[i] if i < len(raw_cfg_disc) else {}, disc, {p['nome'] for p in profs}) for i, disc in enumerate(discs)]
        raw_cfg_prof = data.get('config_professores', [])
        if not isinstance(raw_cfg_prof, list): raise ValueError('Backup inválido: configurações de professores devem formar uma lista.')
        data['config_professores'] = [_sanitize_imported_prof_config(raw_cfg_prof[i] if i < len(raw_cfg_prof) else {}, len(discs)) for i, _prof in enumerate(profs)]
        data['grupos_choque'] = _sanitize_imported_groups(data.get('grupos_choque', []), discs)
        data['config_avancadas'] = solver_normalizar_avancadas(data.get('config_avancadas', {}))
        # Dados leves → sessão
        for k in _SESSION_KEYS:
            if k in data: session[k] = data[k]
        _sincronizar_professores_fixos()
        # Resultados → store (se houver)
        res_keys = ['resultado_grade_disciplinas','resultado_grade_professores',
                    'resultado_horarios_professores','resultado_grade_por_semestre',
                    'resultado_grade_por_professor','resultado_grades_com_prof','resultado_relatorio']
        if any(k in data and data[k] for k in res_keys):
            dados = {
                'grade_disciplinas':     data.get('resultado_grade_disciplinas', {}),
                'grade_professores':     data.get('resultado_grade_professores', {}),
                'horarios_professores':  data.get('resultado_horarios_professores', {}),
                'grade_por_semestre':    data.get('resultado_grade_por_semestre', {}),
                'grade_por_professor':   data.get('resultado_grade_por_professor', {}),
                'grades_com_prof':        data.get('resultado_grades_com_prof', {}),
                'relatorio':             data.get('resultado_relatorio', {}),
            }
            token = _store_resultados(dados)
            session['resultado_token'] = token
        session.modified = True
        auto_save()
        return jsonify({'status':'sucesso'})
    except (ValueError, json.JSONDecodeError, UnicodeDecodeError) as exc:
        return jsonify({'status':'erro','mensagem':'Backup inválido: '+str(exc)}), 400
    except Exception:
        logging.exception('[Combinix] Erro ao importar backup')
        return jsonify({'status':'erro','mensagem':'Não foi possível importar o backup.'}), 500

@app.route('/reset', methods=['POST'])
def reset():
    """Reset COMPLETO: limpa sessão + state.json + resultados do usuário."""
    workspace = _workspace_id()
    _limpar_resultados_usuario()
    session.clear()
    if STORAGE_MODE == 'local':
        reset_state(workspace)
    return jsonify({'status':'ok'})

@app.route('/reset_configuracoes', methods=['POST'])
def reset_configuracoes():
    """
    Reset SUAVE: mantém seleções (disciplinas + professores escolhidos),
    mas zera todas as configurações (fixações, tipos, conflitos, indisponibilidades).
    """
    discs = session.get('disciplinas_selecionadas', [])
    profs = session.get('professores_selecionados', [])
    session['config_disciplinas']  = [_cfg_padrao(d) for d in discs]
    session['config_professores']  = [_cfg_padrao_prof() for _ in profs]
    session['grupos_choque']       = []
    session['config_avancadas']    = {'estado_sabado':'desativado','nivel_restricao':3}
    # Limpa resultados do usuário (store + token)
    _limpar_resultados_usuario()
    session.modified = True
    auto_save()
    return jsonify({'status':'ok'})

@app.route('/reset_resultados', methods=['POST'])
def reset_resultados():
    """Reset apenas dos resultados gerados — mantém TUDO (seleções + configs)."""
    _limpar_resultados_usuario()
    session.modified = True
    auto_save()
    return jsonify({'status':'ok'})

@app.route('/api/storage/status')
def api_storage_runtime_status():
    return jsonify({
        'mode': STORAGE_MODE,
        'browser_primary': 'IndexedDB' if STORAGE_MODE == 'browser' else None,
        'browser_fallback': 'localStorage' if STORAGE_MODE == 'browser' else None,
        'message': ('Dados armazenados neste navegador.' if STORAGE_MODE == 'browser' else 'Dados armazenados na pasta local do Combinix.'),
    })


@app.route('/salvar_tema', methods=['POST'])
def salvar_tema():
    data=request.get_json(force=True,silent=True) or {}
    if not data: data=request.form.to_dict()
    session['tema']=data.get('tema','claro') if data.get('tema','claro') in {'claro','escuro'} else 'claro'; session.modified=True; auto_save()
    return jsonify({'status':'ok'})

if __name__ == '__main__':
    app.run(debug=os.environ.get('COMBINIX_DEBUG') == '1', host=os.environ.get('COMBINIX_HOST', '127.0.0.1'), port=int(os.environ.get('COMBINIX_PORT', '5000')))